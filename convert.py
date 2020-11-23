import pandas
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

wb = load_workbook('data/regions.xlsx')
ws = wb.active
df = pandas.read_excel('data/all_shifts.xlsx')
df1 = df[['Sales Rep', 'Cost per', 'Units Sold']]
df1['Total'] = df1['Cost per'] * df1['Units Sold']

rows = dataframe_to_rows(df1, index=False)

for row_index, row in enumerate(rows, 1):
  for column_index, col in enumerate(row, 6):
    ws.cell(row=row_index, column=column_index, value=col)

wb.save('output/combinded.xlsx')
