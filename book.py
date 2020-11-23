from openpyxl.workbook import Workbook
from openpyxl import load_workbook

workbook = Workbook()
worksheet = workbook.active

ws1 = workbook.create_sheet('New Shiiit')
ws2 = workbook.create_sheet('Another sheet', 0)

worksheet.title = 'mySheet'

workbook2 = load_workbook('data/regions.xlsx')
new_sheet = workbook2.create_sheet('yet another sheet')
active_sheet = workbook2.active
cell = active_sheet['A1']

active_sheet['A1'] = f'helllllo'
workbook2.save('output/modified.xlsx')
