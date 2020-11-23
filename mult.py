import pandas
from openpyxl import load_workbook
from openpyxl.styles import Font

df1 = pandas.read_excel('data/shifts.xlsx', sheet_name='Sheet')
df2 = pandas.read_excel('data/shifts.xlsx', sheet_name='Sheet1')
df3 = pandas.read_excel('data/shift_3.xlsx')

df_all = pandas.concat([df1, df2, df3], sort=False)

to_excel = df_all.to_excel('output/allshifts.xlsx', index=None)

wb = load_workbook('output/allshifts.xlsx')
ws = wb.active

total_col = ws['G1']
total_col.font = Font(bold=True)
total_col.value = 'Total'

e_col, f_col = ['E', 'F']
for row in range(2, 300):
  result_cell = 'G{}'.format(row)
  e_value = ws[e_col + str(row)].value
  f_value = ws[f_col + str(row)].value
  ws[result_cell] = e_value * f_value

wb.save('output/totalled.xlsx')
