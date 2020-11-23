import pandas, numpy
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

wb = load_workbook('data/template.xlsx')
ws = wb.active

df = pandas.read_csv('data/crime.csv', encoding='utf-8', dtype={"INCIDENT_NUMBER": str, "OFFENSE_CODE": str, "OFFENSE_CODE_GROUP": str, "OFFENSE_DESCRIPTION": str, "DISTRICT": str, "REPORTING_AREA": str, "SHOOTING": str, "YEAR": str, "MONTH": str, "DAY_OF_WEEK": str, "HOUR": str})

df1 = df[df['OFFENSE_CODE_GROUP'] == 'Counterfeiting']
df1 = df1.replace(numpy.nan, 'N/A', regex=True)

total_crimes = len(df.index)
counterfeit = len(df1.index)
perc_crimes = (counterfeit / total_crimes) * 100
perc_crimes = round(perc_crimes, 2)

ws['O8'].value = total_crimes
ws['P8'].value = counterfeit
ws['Q8'].value = perc_crimes

df1['Count'] = 1
df2 = df1.groupby(['DISTRICT', 'YEAR']).count()['Count'].unstack(level=0)
df2.drop(columns='N/A', inplace=True)

rows = dataframe_to_rows(df2)
for row_index, row in enumerate(rows, 8):
  for column_index, value in enumerate(row, 1):
    ws.cell(row=row_index, column=column_index, value=value)

wb.save('output/crime_report.xlsx')
